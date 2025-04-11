import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import datetime
import matplotlib.ticker as mtick
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def format_data(path):
    df = pd.read_csv(path)
    df.columns = df.columns.str.strip()  # Removes leading/trailing spaces
    df.drop(columns = ["Veterinary Sciences Library"], inplace= True)


    col_list = list(df.columns)
    x, y = col_list.index("Grace Reeves Study Centre"), col_list.index("Hawthorns & Brambles")
    col_list[y], col_list[x] = col_list[x], col_list[y]
    df = df[col_list]
    

    # Rename columns so libraries have the same names as in library_locations.csv
    library_data = pd.read_csv(r"C:\Users\acdal\OneDrive - University of Bristol\2024-2025\MDM3\Phase C\MDM3-phaseC\library_locations.csv")
    library_names = library_data['LibraryName '].tolist()
    df.columns = list(df.columns[:4]) + library_names

    df_values = df.loc[:, 'Arts and Social Sciences':]
    df_values = df_values.apply(pd.to_numeric, errors='coerce')


    capacity = library_data["Capacity"].astype(int).to_numpy()
    df_values_percentage = df_values.div(capacity,axis = 1) * 100

    df_percentage_formatted = pd.concat([df.iloc[:,:4],df_values_percentage],axis = 1)

    return df_percentage_formatted


def week_formatting(df):

    # Remove '00:00:00' from all elements in 'W/C' col
    df['W/C'] = df['W/C'].to_frame().applymap(lambda x: str(x).replace(' 00:00:00', '').strip() if isinstance(x, str) else x)
    weeks = df['W/C'].unique()

    df.to_csv("week_formatted_percentage_library_occupancy_23_24.csv", index = False)

    return weeks


def weekly_percentage_full_all_libraries(df, upper_threshold = None, lower_threshold = None):
    
    weeks = df['W/C'].unique()

    weekly_percentage_df = pd.DataFrame(weeks, columns = ['W/C'])

    # Ensure the first column is in datetime format
    weekly_percentage_df['W/C'] = pd.to_datetime(weekly_percentage_df['W/C'])    
    weekly_percentage_df['W/C'] = weekly_percentage_df['W/C'].dt.strftime('%d-%m-%Y')

    if upper_threshold != None:
        for threshold in upper_threshold:
            weekly_percentage_df[f'{threshold}% full'] = np.nan

            for week in enumerate(weeks):
                
                current_week_data = df[df['W/C'] == week[1]]
                current_week_occupancy_values = current_week_data.loc[:,'Arts and Social Sciences':]

                # Flatten all values into a single Series and drop NaNs
                all_values = current_week_occupancy_values.values.flatten()
                all_values = pd.Series(all_values).dropna()

                
                # Compute proportion of values ≥ upper_threshold across all libraries
                overall_proportion = (all_values >= threshold).sum() / all_values.count()
                weekly_percentage_df.at[week[0],f'{threshold}% full'] = overall_proportion * 100

                # elif lower_threshold != None:
                #     # Compute proportion of values ≤ lower_threshold across all libraries
                #     overall_proportion = (all_values <= threshold).sum() / all_values.count()
                #     weekly_percentage_df.at[week[0],f'{threshold}% full'] = overall_proportion * 100

    
        # Plot proportion of time library occupancy exceeds threshold per academic week
        
        fig, ax = plt.subplots(figsize=(10, 6))
        
        legend_thresholds = weekly_percentage_df.columns[1:]
        print(weekly_percentage_df.dtypes)

        test = weekly_percentage_df.at[16,'W/C']

        # Plot each threshold on one figure
        for threshold in legend_thresholds:
            ax.plot(weekly_percentage_df['W/C'], weekly_percentage_df[threshold], marker='o', linestyle='-',label = threshold)
        
        

        ax.axvline(x=weekly_percentage_df.at[16,'W/C'], linestyle='--', linewidth = 1.5, color = 'tab:purple', label='Assessment period')
        ax.axvline(x=weekly_percentage_df.at[18,'W/C'], linestyle='--', linewidth = 1.5, color = 'tab:purple')    

        ax.axvline(x=weekly_percentage_df.at[34,'W/C'], linestyle='--', linewidth = 1.5, color = 'tab:purple')
        ax.axvline(x=weekly_percentage_df.at[37,'W/C'], linestyle='--', linewidth = 1.5, color = 'tab:purple')   

        ax.axvline(x=weekly_percentage_df.at[6,'W/C'], linestyle='--', linewidth = 1.5, color = 'tab:gray', label='Reading week')
        ax.axvline(x=weekly_percentage_df.at[7,'W/C'], linestyle='--', linewidth = 1.5, color = 'tab:gray')

        ax.axvline(x=weekly_percentage_df.at[23,'W/C'], linestyle='--', linewidth = 1.5, color = 'tab:gray')
        ax.axvline(x=weekly_percentage_df.at[24,'W/C'], linestyle='--', linewidth = 1.5, color = 'tab:gray')   

        ax.yaxis.set_major_formatter(mtick.PercentFormatter())
        ax.set_ylim([0,100])
        ax.set_yticks(np.arange(0, 101, 10))  # From 0% to 100% in steps of 10%
        ax.legend(title = 'Threshold occupancy')
        
        # Rotate x-axis labels for better readability
        plt.xticks(rotation=45)  

        # Labels and title
        plt.xlabel("Week",fontsize = 12)
        plt.ylabel("Proportion of time libraries exceed threshold occupancy", fontsize = 12)
        # # plt.title("Proportion of time library occupancy exceeds threshold per academic week")

        plt.tight_layout()
        plt.grid()
        plt.show()

    elif lower_threshold != None:
        for threshold in lower_threshold:
            weekly_percentage_df[f'{threshold}% full'] = np.nan

            for week in enumerate(weeks):
                
                current_week_data = df[df['W/C'] == week[1]]
                current_week_occupancy_values = current_week_data.loc[:,'Arts and Social Sciences':]

                # Flatten all values into a single Series and drop NaNs
                all_values = current_week_occupancy_values.values.flatten()
                all_values = pd.Series(all_values).dropna()

                
                # Compute proportion of values ≥ upper_threshold across all libraries
                overall_proportion = (all_values <= threshold).sum() / all_values.count()
                weekly_percentage_df.at[week[0],f'{threshold}% full'] = overall_proportion * 100

    
        # Plot proportion of time library occupancy exceeds threshold per academic week
        
        fig, ax = plt.subplots(figsize=(10, 6))
        
        legend_thresholds = weekly_percentage_df.columns[1:]
        print(weekly_percentage_df.dtypes)

        test = weekly_percentage_df.at[16,'W/C']

        # Plot each threshold on one figure
        for threshold in legend_thresholds:
            ax.plot(weekly_percentage_df['W/C'], weekly_percentage_df[threshold], marker='o', linestyle='-',label = threshold)
        
        

        ax.axvline(x=weekly_percentage_df.at[16,'W/C'], linestyle='--', linewidth = 1.5, color = 'tab:purple', label='Assessment period')
        ax.axvline(x=weekly_percentage_df.at[18,'W/C'], linestyle='--', linewidth = 1.5, color = 'tab:purple')    

        ax.axvline(x=weekly_percentage_df.at[34,'W/C'], linestyle='--', linewidth = 1.5, color = 'tab:purple')
        ax.axvline(x=weekly_percentage_df.at[37,'W/C'], linestyle='--', linewidth = 1.5, color = 'tab:purple')   

        ax.axvline(x=weekly_percentage_df.at[6,'W/C'], linestyle='--', linewidth = 1.5, color = 'tab:gray', label='Reading week')
        ax.axvline(x=weekly_percentage_df.at[7,'W/C'], linestyle='--', linewidth = 1.5, color = 'tab:gray')

        ax.axvline(x=weekly_percentage_df.at[23,'W/C'], linestyle='--', linewidth = 1.5, color = 'tab:gray')
        ax.axvline(x=weekly_percentage_df.at[24,'W/C'], linestyle='--', linewidth = 1.5, color = 'tab:gray')   

        ax.yaxis.set_major_formatter(mtick.PercentFormatter())
        ax.set_ylim([0,100])
        ax.set_yticks(np.arange(0, 101, 10))  # From 0% to 100% in steps of 10%
        ax.legend(title = 'Threshold occupancy')
        
        # Rotate x-axis labels for better readability
        plt.xticks(rotation=45)  

        # Labels and title
        plt.xlabel("Week",fontsize = 12)
        plt.ylabel("Proportion of time libraries are below threshold occupancy", fontsize = 12)
        # # plt.title("Proportion of time library occupancy exceeds threshold per academic week")

        plt.tight_layout()
        plt.grid()
        plt.show()


    return weekly_percentage_df


def hourly_percentage_full_all_libraries(df, upper_threshold = None, lower_threshold = None):
    
    time_range = df['Time range'].unique()

    hourly_percentage_df = pd.DataFrame(time_range, columns = ['Time range'])

    print(hourly_percentage_df.dtypes)

    hourly_percentage_df.sort_values(by = ['Time range'], axis = 0, inplace=True)

    # # **Ensure column is string and remove spaces**
    # hourly_percentage_df.iloc[:, 0] = hourly_percentage_df.iloc[:, 0].astype(str).str.strip()

    # # **Use regex to extract only the first HH:MM part**
    # hourly_percentage_df.iloc[:, 0] = hourly_percentage_df.iloc[:, 0].apply(lambda x: re.match(r"(\d{2}:\d{2})", x).group(1) if re.match(r"(\d{2}:\d{2})", x) else x)

    # # Convert to datetime format for proper sorting
    # hourly_percentage_df.iloc[:, 0] = pd.to_datetime(hourly_percentage_df.iloc[:, 0], format="%H:%M")

    # # Sort by time
    # hourly_percentage_df = hourly_percentage_df.sort_values(by=hourly_percentage_df.columns[0])

    # hourly_percentage_df['Time range'] = pd.to_datetime(hourly_percentage_df['Time range'])
    # hourly_percentage_df["Time range"] = hourly_percentage_df["Time range"].dt.strftime("%H:%M")

    # # Reset index
    # hourly_percentage_df = hourly_percentage_df.reset_index(drop=True)

    # Ensure the first column is in datetime format
    # hourly_percentage_df['Time range'] = pd.to_datetime(hourly_percentage_df['Time range'])    
    # hourly_percentage_df['Time range'] = hourly_percentage_df['Time range'].dt.strftime('%d-%m-%Y')

    if upper_threshold != None:
        for threshold in upper_threshold:
            hourly_percentage_df[f'{threshold}% full'] = np.nan

            for time in enumerate(time_range):
                
                current_hour_data = df[df['Time range'] == time[1]]
                current_hour_occupancy_values = current_hour_data.loc[:,'Arts and Social Sciences':]

                # Flatten all values into a single Series and drop NaNs
                all_values = current_hour_occupancy_values.values.flatten()
                all_values = pd.Series(all_values).dropna()

                
                # Compute proportion of values ≥ upper_threshold across all libraries
                overall_proportion = (all_values >= threshold).sum() / all_values.count()
                hourly_percentage_df.at[time[0],f'{threshold}% full'] = overall_proportion * 100    

        
        # Plot proportion of the time that library occupancy exceeds the threshold per hour
        
        fig, ax = plt.subplots(figsize=(10, 6))
        
        legend_thresholds = hourly_percentage_df.columns[1:]
        print(hourly_percentage_df.dtypes)

        # test = weekly_percentage_df.at[16,'W/C']

        # Plot each threshold on one figure
        for threshold in legend_thresholds:
            ax.plot(hourly_percentage_df['Time range'], hourly_percentage_df[threshold], marker='o', linestyle='-',label = threshold)

        ax.yaxis.set_major_formatter(mtick.PercentFormatter())
        ax.set_ylim([0,100])
        ax.set_yticks(np.arange(0, 101, 10))  # From 0% to 100% in steps of 10%
        ax.legend(title = 'Threshold occupancy')
        
        # Rotate x-axis labels for better readability
        plt.xticks(rotation=75)  

        # Labels and title
        plt.xlabel("Time of Day",fontsize = 12)
        plt.ylabel("Proportion of time libraries are above threshold occupancy", fontsize = 12)
        # # plt.title("Proportion of time library occupancy exceeds threshold per academic week")

        plt.tight_layout()
        plt.grid()
        plt.show()


    return hourly_percentage_df


def remove_weekend(df):

    df = df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
    # Remove data from Saturday and Sunday
    # df['Day'] = df['Day'].str.strip()
    df_filtered = df.drop(df[(df['Day'] == 'Saturday') | (df['Day'] == 'Sunday')].index) 
    return df_filtered


def identify_peak_hour(df_filtered, upper_threshold):

    
    # Extract data for Arts and Social Sciences Library
    df_filtered_ASS = df_filtered.loc[:,:'Arts and Social Sciences']

    # Create dataframe to compare mean occupancy and proportion of time above threshold occupancy
    times = df_filtered_ASS['Time range'].unique()
    times_df = pd.DataFrame(times,columns = ['Time range'])
    times_df.sort_values(by = ['Time range'], axis = 0, inplace=True)
    
    times_df['Mean occupancy'] = np.nan
    times_df['Median occupancy'] = np.nan
    times_df['Proportion above threshold occupancy'] = np.nan
    

    for time in enumerate(times):
        current_time_data = df_filtered_ASS[df_filtered_ASS['Time range'] == time[1]]
        mean_occupancy = current_time_data['Arts and Social Sciences'].mean()
        times_df.at[time[0],'Mean occupancy'] = mean_occupancy

        median_occupancy = current_time_data['Arts and Social Sciences'].median()
        times_df.at[time[0],'Median occupancy'] = median_occupancy

        ASS_occupancy = current_time_data['Arts and Social Sciences'].dropna()

        # Compute proportion of values ≥ upper_threshold in the ASS library
        proportion = (ASS_occupancy >= upper_threshold).sum() / ASS_occupancy.count()
        times_df.at[time[0],'Proportion above threshold occupancy'] = proportion * 100
    
    
    # times_df.iloc[:, 0] = times_df.iloc[:, 0].apply(lambda x: re.match(r"(\d{2}:\d{2})", x).group(1) if re.match(r"(\d{2}:\d{2})", x) else x)

    # # Convert to datetime format for proper sorting
    # times_df['Time range'] = pd.to_datetime(times_df.iloc[:, 0], format="%H:%M")
    # times_df["Time range"] = times_df["Time range"].dt.strftime("%H:%M")
    
    
    times_df = times_df.reset_index(drop = True)
    


    

    # # Plot 
    # fig, ax = plt.subplots(figsize=(10, 6))
        
    # ax.plot(times_df['Time range'], times_df['Median occupancy'], marker='o', linestyle='-',label = 'Arts and Social Sciences')

    # ax.yaxis.set_major_formatter(mtick.PercentFormatter())
    # ax.set_ylim([0,100])
    # # ax.set_xlim([times_df['Time range'][0],times_df['Time range'][23]])
    # ax.set_yticks(np.arange(0, 101, 10))  # From 0% to 100% in steps of 10%
    # ax.legend(title = 'Library')
        
    # # Rotate x-axis labels for better readability
    # plt.xticks(rotation=45)  

    # # Labels and title
    # plt.xlabel("Hour of the Day",fontsize = 14)
    # plt.ylabel("Median occupancy", fontsize = 14)

    # plt.tight_layout()
    # plt.grid()
    # plt.show()


    times_df['Average metric score'] = np.nan
    # test = times_df.loc[:,'Mean occupancy':].mean(axis = 1)
    times_df['Average metric score'] = times_df.loc[:,'Mean occupancy':].mean(axis = 1)

    return times_df


def ASS_comparison(df,upper_threshold, other_library):

    df_filtered = remove_weekend(df)

    times_df = identify_peak_hour(df_filtered,upper_threshold)

    # Extract peak time based on mean occupancy, meadian occupancy, and proportion of time above threshold occupancy (e.g. 90%)
    peak_time = times_df['Time range'][times_df['Average metric score'].idxmax()]

    # peak_time = '12:00-12:59'
    # peak_time = '13:00-13:59'
    peak_time_df = df_filtered[df_filtered['Time range'] == peak_time]
    # peak_time_df = df_filtered[(df_filtered['Time range'] == '12:00-12:59') | (df_filtered['Time range'] == '13:00-13:59') | (df_filtered['Time range'] == '15:00-15:59')]
    # peak_time_df = df_filtered[(df_filtered['Time range'] == '13:00-13:59') | (df_filtered['Time range'] == '15:00-15:59')]

    # Extract peak time data for ASS and another library (Physics library has a lot of data for 15:00-15:59)
    ASS_other_peak_time_df = pd.concat([peak_time_df.loc[:,:'Arts and Social Sciences'], peak_time_df[other_library]], axis = 1)

    # ASS_other_peak_time_df_cleaned = ASS_other_peak_time_df.dropna(subset=['Arts and Social Sciences',other_library])

    # ASS_physics_peak_time_df_cleaned = ASS_physics_peak_time_df.dropna(subset=['Arts and Social Sciences','Physics'], how = 'all')

    # weeks = ASS_other_peak_time_df_cleaned['W/C'].unique()
    weeks = ASS_other_peak_time_df['W/C'].unique()
    weeks_df = pd.DataFrame(weeks,columns=['W/C'])
    weeks_df['ASS median occupancy'] = np.nan

    weeks_df[f'{other_library} median occupancy'] = np.nan

    # Ensure the first column is in datetime format
    weeks_df['W/C'] = pd.to_datetime(weeks_df['W/C'])    
    weeks_df['W/C'] = weeks_df['W/C'].dt.strftime('%d-%m-%Y')


    for week in enumerate(weeks):
        # current_week_data = ASS_other_peak_time_df_cleaned[ASS_other_peak_time_df_cleaned['W/C'] == week[1]]
        current_week_data = ASS_other_peak_time_df[ASS_other_peak_time_df['W/C'] == week[1]]
        ASS_median = current_week_data['Arts and Social Sciences'].median()
        weeks_df.at[week[0],'ASS median occupancy'] = ASS_median

        other_median = current_week_data[other_library].median()
        weeks_df.at[week[0],f'{other_library} median occupancy'] = other_median

    weeks_df_cleaned = weeks_df.dropna(subset=['ASS median occupancy',f'{other_library} median occupancy'])
    weeks_df_cleaned = weeks_df_cleaned.reset_index(drop=True)

    # # # For Physics
    # weeks_df_cleaned.drop(index=[0,1,2,3,10],inplace=True)
    # # # weeks_df_cleaned.drop(index=[0,1,2,3,4,5,12,13],inplace=True)

    # # # For Wills Memorial
    # # # weeks_df_cleaned.drop(index=[6],inplace=True)

    # # Plot 
    # fig, ax = plt.subplots(figsize=(10, 6))
        
    # ax.plot(weeks_df_cleaned['W/C'], weeks_df_cleaned['ASS median occupancy'], marker='o', linestyle='-',label = 'Arts and Social Sciences')
    # ax.plot(weeks_df_cleaned['W/C'], weeks_df_cleaned[f'{other_library} median occupancy'], marker='o', linestyle='-',label = other_library)

    # ax.yaxis.set_major_formatter(mtick.PercentFormatter())
    # ax.set_ylim([0,100])
    # ax.set_yticks(np.arange(0, 101, 10))  # From 0% to 100% in steps of 10%
    # ax.legend(title = 'Library')
        
    # # Rotate x-axis labels for better readability
    # # plt.xticks(rotation=30)  

    # # Labels and title
    # plt.xlabel("Week",fontsize = 14)
    # plt.ylabel("Median occupancy during peak hour", fontsize = 14)

    # plt.tight_layout()
    # plt.grid()
    # plt.show()

    return weeks_df_cleaned





if __name__ == "__main__":

    # # Load the Excel file
    # file_path = r"C:\Users\acdal\OneDrive - University of Bristol\2024-2025\MDM3\Phase C\MDM3-phaseC\FOR_DISPLAY_percentage_library_occupancy_23_24.xlsx"
    # df = pd.read_excel(file_path)

    # # Load workbook and select active worksheet
    # wb = load_workbook(file_path)
    # ws = wb.active  

    # # Define red fill style
    # red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # # Identify column indices (Excel columns are 1-based, pandas are 0-based)
    # col_start, col_end = 5, 15  # Adjusted for Excel's 1-based index

    # # Loop through the specified columns and apply red fill to empty cells
    # for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_start, max_col=col_end):
    #     for cell in row:
    #         if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ""):
    #             cell.fill = red_fill  # Apply red background

    # # Save the modified file
    # wb.save("highlighted_missing_data.xlsx")

    # print("Excel formatting done!")






    path = r"C:\Users\acdal\OneDrive - University of Bristol\2024-2025\MDM3\Phase C\MDM3-phaseC\cleaned_library_occupancy_data_23_24.csv"
    
    df_percentage_formatted = format_data(path)

    df_percentage_formatted.to_csv("percentage_library_occupancy_23_24.csv", index=False)

    # df_percentage_formatted = pd.read_csv(r"C:\Users\acdal\OneDrive - University of Bristol\2024-2025\MDM3\Phase C\MDM3-phaseC\percentage_library_occupancy_23_24.csv")    

    # weeks = week_formatting(df_percentage_formatted)

    # df = pd.read_csv(r"C:\Users\acdal\OneDrive - University of Bristol\2024-2025\MDM3\Phase C\MDM3-phaseC\week_formatted_percentage_library_occupancy_23_24.csv")
    df = pd.read_csv(r"C:\Users\acdal\OneDrive - University of Bristol\2024-2025\MDM3\Phase C\MDM3-phaseC\percentage_library_occupancy_23_24.csv")

    # weekly_occupancy = weekly_percentage_full_all_libraries(df, upper_threshold = [25,50,75,95], lower_threshold = [10,25])

    # hourly_occupancy = hourly_percentage_full_all_libraries(df, upper_threshold = [25,60,75,95], lower_threshold = [10,25])

    # individual_occupancy = individual_library_fullness(df, upper_threshold = [25,50,75, 95], lower_threshold = [10,25])

    physics_peak_hour_comparison = ASS_comparison(df,upper_threshold=90, other_library="Physics")

    wills_memorial_peak_hour_comparison = ASS_comparison(df, upper_threshold=90, other_library="Wills Memorial")


    # For Physics
    physics_peak_hour_comparison.drop(index=[0,1,2,3,10],inplace=True)
    # weeks_df_cleaned.drop(index=[0,1,2,3,4,5,12,13],inplace=True)

    # For Wills Memorial
    wills_memorial_peak_hour_comparison.drop(index=[6],inplace=True)

    # Plot 
    fig, ax = plt.subplots(figsize=(10, 6))
        
    ax.plot(physics_peak_hour_comparison['W/C'], physics_peak_hour_comparison['ASS median occupancy'], marker='o', linestyle='-',label = 'Arts and Social Sciences')
    
    ax.plot(wills_memorial_peak_hour_comparison['W/C'], wills_memorial_peak_hour_comparison['Wills Memorial median occupancy'], marker='o', linestyle='-',label = 'Wills Memorial')
    ax.plot(physics_peak_hour_comparison['W/C'], physics_peak_hour_comparison['Physics median occupancy'], marker='o', linestyle='-',label = 'Physics')
    
    ax.yaxis.set_major_formatter(mtick.PercentFormatter())
    ax.set_ylim([0,100])
    ax.set_yticks(np.arange(0, 101, 10))  # From 0% to 100% in steps of 10%
    ax.legend(title = 'Library')
        
    # Rotate x-axis labels for better readability
    # plt.xticks(rotation=30)  

    # Labels and title
    plt.xlabel("Week",fontsize = 14)
    plt.ylabel("Median occupancy during peak hour", fontsize = 14)

    plt.tight_layout()
    plt.grid()
    plt.show()



    print("DONE!")




    # def individual_library_fullness(df, upper_threshold, lower_threshold):

#     # time_range = df['Time range'].unique()

#     # hourly_percentage_df = pd.DataFrame(time_range, columns = ['Time range'])

#     # hourly_percentage_df.sort_values(by = ['Time range'], axis = 0, inplace=True)

#     # # **Ensure column is string and remove spaces**
#     # hourly_percentage_df.iloc[:, 0] = hourly_percentage_df.iloc[:, 0].astype(str).str.strip()

#     # # **Use regex to extract only the first HH:MM part**
#     # hourly_percentage_df.iloc[:, 0] = hourly_percentage_df.iloc[:, 0].apply(lambda x: re.match(r"(\d{2}:\d{2})", x).group(1) if re.match(r"(\d{2}:\d{2})", x) else x)

#     # # Convert to datetime format for proper sorting
#     # hourly_percentage_df.iloc[:, 0] = pd.to_datetime(hourly_percentage_df.iloc[:, 0], format="%H:%M")

#     # # Sort by time
#     # hourly_percentage_df = hourly_percentage_df.sort_values(by=hourly_percentage_df.columns[0])

#     # hourly_percentage_df['Time range'] = pd.to_datetime(hourly_percentage_df['Time range'])
#     # hourly_percentage_df["Time range"] = hourly_percentage_df["Time range"].dt.strftime("%H:%M")

#     # # Reset index
#     # hourly_percentage_df = hourly_percentage_df.reset_index(drop=True)

#     df["Time range"] = df["Time range"].str.extract(r"(\d{2}:\d{2})")
#     df["Time range"] = pd.to_datetime(df["Time range"], format="%H:%M").dt.hour

    
#     return df


# def individuaL_weekly_percentage_full(df, upper_threshold = None, lower_threshold = None):
#     # ASS_library_data = df.loc[:,:'Arts and Social Sciences']
    
#     # weeks = ASS_library_data['W/C'].unique()
#     weeks = df['W/C'].unique()

#     df.loc[:].columns

#     # Keys : Weeks of the academic year 
#     # Values : percent of the time that the library occupancy is above upper threshold
#     weekly_percentage_dict = {}

#     for week in weeks:
        
#         current_week_data = df[df['W/C'] == week]
#         current_week_occupancy_values = current_week_data.loc[:,'Arts and Social Sciences':]

#         # Flatten all values into a single Series and drop NaNs
#         all_values = current_week_occupancy_values.values.flatten()
#         all_values = pd.Series(all_values).dropna()

#         # Compute proportion of values ≥ upper_threshold across all libraries
#         overall_proportion = (all_values >= upper_threshold).sum() / all_values.count()
#         weekly_percentage_dict[week] = overall_proportion * 100

#     # Function to calculate the proportion of values ≥ 95
#     def proportion_ge_95(col):
#         return (col >= 95).sum() / col.notna().sum()

#     # Apply function to each numerical column
#     proportions = df.apply(proportion_ge_95)

#     # for week in weeks:
#     #     current_week_data = ASS_library_data[ASS_library_data['W/C'] == week]

#     #     # Select the column (replace 'Column1' with your actual column name)
#     #     col = current_week_data['Arts and Social Sciences']

#     #     # Calculate the proportion of non-missing values that are ≥ 95
#     #     proportion = (col >= upper_threshold).sum() / col.notna().sum()
#     #     weekly_percentage_dict[week] = proportion * 100


#     return weekly_percentage_dict



















def ASS_comparison(df,upper_threshold):

    df_filtered = remove_weekend(df)

    times_df = identify_peak_hour(df_filtered,upper_threshold)

    # Extract peak time based on mean occupancy, meadian occupancy, and proportion of time above threshold occupancy (e.g. 90%)
    peak_time = times_df['Time range'][times_df['Average metric score'].idxmax()]

    # peak_time = '12:00-12:59'
    # peak_time = '13:00-13:59'
    peak_time_df = df_filtered[df_filtered['Time range'] == peak_time]

    # Extract peak time data for ASS and another library (Physics library has a lot of data for 15:00-15:59)
    ASS_physics_peak_time_df = pd.concat([peak_time_df.loc[:,:'Arts and Social Sciences'], peak_time_df['Physics']], axis = 1)

    ASS_physics_peak_time_df_cleaned = ASS_physics_peak_time_df.dropna(subset=['Arts and Social Sciences','Physics'])
    # ASS_physics_peak_time_df_cleaned = ASS_physics_peak_time_df.dropna(subset=['Arts and Social Sciences','Physics'], how = 'all')

    weeks = ASS_physics_peak_time_df_cleaned['W/C'].unique()
    weeks_df = pd.DataFrame(weeks,columns=['W/C'])
    weeks_df['ASS median occupancy'] = np.nan
    weeks_df['Physics median occupancy'] = np.nan

    # Ensure the first column is in datetime format
    weeks_df['W/C'] = pd.to_datetime(weeks_df['W/C'])    
    weeks_df['W/C'] = weeks_df['W/C'].dt.strftime('%d-%m-%Y')


    for week in enumerate(weeks):
        current_week_data = ASS_physics_peak_time_df_cleaned[ASS_physics_peak_time_df_cleaned['W/C'] == week[1]]
        ASS_median = current_week_data['Arts and Social Sciences'].median()
        weeks_df.at[week[0],'ASS median occupancy'] = ASS_median

        Physics_median = current_week_data['Physics'].median()
        weeks_df.at[week[0],'Physics median occupancy'] = Physics_median

    weeks_df.drop(index=[0,1,2,9],inplace=True)

    # Plot 
    fig, ax = plt.subplots(figsize=(10, 6))
        
    ax.plot(weeks_df['W/C'], weeks_df['ASS median occupancy'], marker='o', linestyle='-',label = 'Arts and Social Sciences')
    ax.plot(weeks_df['W/C'], weeks_df['Physics median occupancy'], marker='o', linestyle='-',label = 'Physics')

    ax.yaxis.set_major_formatter(mtick.PercentFormatter())
    ax.set_ylim([0,100])
    ax.set_yticks(np.arange(0, 101, 10))  # From 0% to 100% in steps of 10%
    ax.legend(title = 'Library')
        
    # Rotate x-axis labels for better readability
    plt.xticks(rotation=30)  

    # Labels and title
    plt.xlabel("Week",fontsize = 12)
    plt.ylabel("Median occupancy during peak hour", fontsize = 12)

    plt.tight_layout()
    plt.grid()
    plt.show()



    return weeks_df
